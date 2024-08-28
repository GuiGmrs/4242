from flask import Flask, request, render_template, redirect
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)

# Caminho para o arquivo Excel
excel_file = "database.xlsx"

# Verifica se o arquivo Excel existe, se não, cria um novo
if not os.path.exists(excel_file):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Produção"
    # Cabeçalhos
    headers = ["Data", "Produto", "Matéria-Prima", "Quantidade de Matéria-Prima", "Quantidade de Produto Final", "Observações"]
    sheet.append(headers)
    workbook.save(excel_file)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/registrar', methods=['POST'])
def registrar():
    # Recebe os dados do formulário
    produto = request.form['produto']
    materia_prima = request.form['materiaPrima']
    qtd_materia_prima = f"{request.form['qtdMateriaPrima']} {request.form['unidadeMateriaPrima']}"
    qtd_produto_final = f"{request.form['qtdProdutoFinal']} {request.form['unidadeProdutoFinal']}"
    observacoes = request.form['observacoes']
    
    # Adiciona os dados ao arquivo Excel
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    data_atual = request.form['dataAtual'] if 'dataAtual' in request.form else None

    # Linha de dados
    dados = [data_atual, produto, materia_prima, qtd_materia_prima, qtd_produto_final, observacoes]
    sheet.append(dados)

    # Salva o arquivo Excel
    workbook.save(excel_file)
    
    return redirect('/')

if __name__ == "__main__":
    app.run(debug=True)
